import sys
import serial
import time
import openpyxl
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel, QVBoxLayout, QHBoxLayout, QWidget, QComboBox, QFileDialog, QMessageBox)

class Worker(QThread):
    update_position = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, serial_port, excel_file, selected_layer):
        super().__init__()
        self.serial_port = serial_port
        self.excel_file = excel_file
        self.selected_layer = selected_layer
        self._is_running = True

    def run(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = workbook[self.selected_layer]
            for row in range(1, 21):
                for col in range(1, 21):
                    if not self._is_running:
                        return
                    cell = sheet.cell(row, col)
                    if cell.fill.start_color.index != '00000000':  # Detects non-default fill
                        command = f"G90\nG0 X{col-1} Y{row-1}\n"
                        self.serial_port.write(command.encode('utf-8'))
                        self.serial_port.flush()
                        time.sleep(1)  # Simulate execution time
                        self.update_position.emit(f"Moved to X{col-1} Y{row-1}")
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))

    def stop(self):
        self._is_running = False

class IndustrialCncGui(QMainWindow):
    
    def __init__(self):
        super().__init__()
        self.serial_port = None
        self.init_ui()
        self.init_serial('/dev/cu.usbserial-10', 115200)  # Update this with your actual port
        
    def init_ui(self):
        self.setWindowTitle('Industrial Paintball CNC Controller')
        self.setGeometry(100, 100, 1280, 720)

        self.load_button = QPushButton('Load Excel File')
        self.layer_combobox = QComboBox()
        self.process_button = QPushButton('Process Layer')
        self.stop_button = QPushButton('STOP')

        # Setting up the layout
        control_layout = QHBoxLayout()
        control_layout.addWidget(self.load_button)
        control_layout.addWidget(self.layer_combobox)
        control_layout.addWidget(self.process_button)
        control_layout.addWidget(self.stop_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(control_layout)

        self.layer_preview = QLabel('Layer Preview')
        self.layer_preview.setStyleSheet("background-color: gray")
        main_layout.addWidget(self.layer_preview)

        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # Connect signals to slots
        self.load_button.clicked.connect(self.load_excel_file)
        self.process_button.clicked.connect(self.process_layer)
        self.stop_button.clicked.connect(self.stop_processing)

    def init_serial(self, port, baud_rate):
        try:
            self.serial_port = serial.Serial(port, baud_rate)
            time.sleep(2)  # Allow time for connection to establish
            print(f"Connected to the device on {port}.")
        except serial.SerialException as e:
            QMessageBox.critical(self, "Serial Connection Error", f"Failed to connect: {e}")

    def load_excel_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)")
        if file_name:
            self.excel_file = file_name
            workbook = openpyxl.load_workbook(file_name)
            sheet_names = workbook.sheetnames
            self.layer_combobox.clear()
            self.layer_combobox.addItems(sheet_names)

    def process_layer(self):
        selected_layer = self.layer_combobox.currentText()
        if not selected_layer or not self.serial_port or not self.serial_port.isOpen():
            QMessageBox.warning(self, "Error", "Please select a layer and ensure the device is connected.")
            return
        
        self.worker = Worker(self.serial_port, self.excel_file, selected_layer)
        self.worker.update_position.connect(lambda msg: print(msg))  # Or update the GUI
        self.worker.finished.connect(lambda: print("Processing finished"))
        self.worker.error.connect(lambda e: QMessageBox.critical(self, "Error", f"An error occurred: {e}"))
        self.worker.start()

    def stop_processing(self):
        if hasattr(self, 'worker') and self.worker.isRunning():
            self.worker.stop()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    gui = IndustrialCncGui()
    gui.show()
    sys.exit(app.exec_())
