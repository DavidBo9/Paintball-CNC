import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel,
                             QVBoxLayout, QHBoxLayout, QWidget, QFrame, QComboBox, QFileDialog, QSlider, QMessageBox, QGridLayout)
from PyQt5.QtCore import Qt, QPropertyAnimation, QRect
from PyQt5.QtGui import QIcon
import openpyxl
import serial 
from openpyxl.styles import PatternFill

class IndustrialCncGui(QMainWindow):
    
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.init_serial('/dev/cu.usbserial-10', 9600)  # Update the port
        
    def init_ui(self):
        # Setup UI elements
        self.setWindowTitle('Industrial Paintball CNC Controller')
        self.setGeometry(100, 100, 1280, 720)
        self.setup_ui_elements()
        self.style_interface()
        
    def init_serial(self, port, baud_rate):
        try:
            self.ser = serial.Serial(port, baud_rate)
        except Exception as e:
            QMessageBox.critical(self, "Serial Connection Error", f"Failed to connect to the Arduino: {e}")
            self.ser = None
 
        
    def setup_ui_elements(self):
        self.setWindowTitle('Industrial Paintball CNC Controller')
        self.setGeometry(100, 100, 1280, 720)

        # Create main layout
        main_layout = QHBoxLayout()

        # Left side layout for loading and processing Excel files
        left_layout = QVBoxLayout()
        self.load_button = QPushButton('Load Excel File')
        self.load_button.setFixedSize(200, 40)
        self.layer_combobox = QComboBox()
        self.layer_preview = QLabel('Layer Preview')
        self.layer_preview.setFixedSize(300, 300)  # Adjust size as needed
        self.layer_preview.setStyleSheet("background-color: gray")
        self.process_button = QPushButton('Process Layer')
        self.process_button.setFixedSize(200, 40)
        self.process_button.setStyleSheet("background-color: green; color: white")
        
        # Add widgets to the left layout
        left_layout.addWidget(self.load_button)
        left_layout.addWidget(self.layer_combobox)
        left_layout.addWidget(self.layer_preview)
        left_layout.addWidget(self.process_button)

        # Right side layout for CNC controls
        right_layout = QVBoxLayout()
        
        self.shoot_button = QPushButton('SHOOT', self)
        self.shoot_button.clicked.connect(lambda: self.send_to_arduino('SHOOT'))
        
        self.home_button = QPushButton('HOME', self)
        self.home_button.clicked.connect(lambda: self.send_to_arduino('HOME'))
        
        self.up_button = QPushButton('UP', self)
        self.up_button.clicked.connect(lambda: self.send_to_arduino('UP'))
        
        self.down_button = QPushButton('DOWN', self)
        self.down_button.clicked.connect(lambda: self.send_to_arduino('DOWN'))
        
        self.left_button = QPushButton('LEFT', self)
        self.left_button.clicked.connect(lambda: self.send_to_arduino('LEFT'))
        
        self.right_button = QPushButton('RIGHT', self)
        self.right_button.clicked.connect(lambda: self.send_to_arduino('RIGHT'))
        
        self.set_x_button = QPushButton('SET X-AXIS', self)
        self.set_x_button.clicked.connect(lambda: self.send_to_arduino('SETX'))
        
        self.set_y_button = QPushButton('SET Y-AXIS', self)
        self.set_y_button.clicked.connect(lambda: self.send_to_arduino('SETY'))
        
        self.reset_button = QPushButton('RESET', self)
        self.reset_button.clicked.connect(lambda: self.send_to_arduino('RESET'))
        
        self.stop_button = QPushButton('STOP', self)
        self.stop_button.setFixedSize(200, 100)
        self.stop_button.setStyleSheet("background-color: red; color: white")
        self.stop_button.clicked.connect(lambda: self.send_to_arduino('STOP'))

        # Create grid layout for CNC movement controls
        move_grid = QGridLayout()
        move_grid.addWidget(self.home_button, 0, 1)
        move_grid.addWidget(self.up_button, 1, 1)
        move_grid.addWidget(self.left_button, 2, 0)
        move_grid.addWidget(self.right_button, 2, 2)
        move_grid.addWidget(self.down_button, 2, 1)

        # Add movement grid and other buttons to the right layout
        right_layout.addLayout(move_grid)
        right_layout.addWidget(self.shoot_button)
        right_layout.addWidget(self.set_x_button)
        right_layout.addWidget(self.set_y_button)
        right_layout.addWidget(self.reset_button)
        right_layout.addWidget(self.stop_button)

        # Add left and right layouts to the main layout
        main_layout.addLayout(left_layout)
        main_layout.addLayout(right_layout)

        # Create a central widget to hold the main layout
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # Connect buttons to their functions
        self.load_button.clicked.connect(self.load_excel_file)
        self.process_button.clicked.connect(self.process_layer)
        # Add connections for movement and action buttons
        self.style_interface()
        
    def create_control_panel(self):
        panel = QFrame()
        panel.setFrameStyle(QFrame.Panel | QFrame.Raised)
        panel_layout = QVBoxLayout()

        # Create buttons and add them to the panel layout
        load_button = QPushButton('Load Excel File')
        load_button.setIcon(QIcon('load_icon.png'))  # Replace with your load icon path
        load_button.clicked.connect(self.load_excel_file)

        process_button = QPushButton('Process Layer')
        process_button.setIcon(QIcon('process_icon.png'))  # Replace with your process icon path
        process_button.clicked.connect(self.process_layer)

        # Add manual movement controls
        self.add_movement_controls(panel_layout)
        

        
        # Add layer preview
        self.layer_preview = QLabel("Layer Preview")
        self.layer_preview.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        panel_layout.addWidget(self.layer_preview)

        # ComboBox for layer selection
        self.layer_combobox = QComboBox()
        self.layer_combobox.addItem("Select Layer")

        # Add widgets to the panel layout
        panel_layout.addWidget(load_button)
        panel_layout.addWidget(self.layer_combobox)
        panel_layout.addWidget(process_button)

        # Add the layout to the panel
        panel.setLayout(panel_layout)

        return panel
    
    def style_interface(self):
        # Style the main window background and other elements
        self.setStyleSheet("""
        QMainWindow {
            background-color: #2c3e50;
        }
        QPushButton {
            background-color: #3498db;
            color: white;
            border-style: outset;
            border-width: 2px;
            border-radius: 10px;
            border-color: beige;
            font: bold 14px;
            padding: 6px;
            min-width: 10em;
            margin: 4px;
        }
        QPushButton#emergencyStopButton {
            background-color: #e74c3c;
        }
        QPushButton#processButton {
            background-color: #2ecc71;
        }
        QLabel {
            background-color: #95a5a6;
            color: white;
            font: bold 14px;
            border-radius: 5px;
            padding: 6px;
            margin: 4px;
        }
        QComboBox {
            background-color: #34495e;
            color: white;
            border-radius: 5px;
            padding: 6px;
            margin: 4px;
            min-width: 10em;
        }
        QSlider::groove:horizontal {
            border: 1px solid #bbb;
            background: white;
            height: 10px;
            border-radius: 4px;
        }
        QSlider::handle:horizontal {
            background: #3498db;
            border: 1px solid #777;
            width: 18px;
            margin: -2px 0;
            border-radius: 4px;
        }
        QSlider::add-page:horizontal {
            background: #bbb;
        }
        QSlider::sub-page:horizontal {
            background: #3498db;
        }
        """)

        # Apply the button styles using objectName to distinguish between different buttons
        self.load_button.setObjectName('loadButton')
        self.process_button.setObjectName('processButton')
        self.shoot_button.setObjectName('shootButton')
        self.home_button.setObjectName('homeButton')
        self.up_button.setObjectName('upButton')
        self.down_button.setObjectName('downButton')
        self.left_button.setObjectName('leftButton')
        self.right_button.setObjectName('rightButton')
        self.set_x_button.setObjectName('setXButton')
        self.set_y_button.setObjectName('setYButton')
        self.reset_button.setObjectName('resetButton')
        self.stop_button.setObjectName('emergencyStopButton')

        # Reapply styles to update the widgets with the new object names
        self.load_button.setStyleSheet(self.styleSheet())
        self.process_button.setStyleSheet(self.styleSheet())
        self.shoot_button.setStyleSheet(self.styleSheet())
        self.home_button.setStyleSheet(self.styleSheet())
        self.up_button.setStyleSheet(self.styleSheet())
        self.down_button.setStyleSheet(self.styleSheet())
        self.left_button.setStyleSheet(self.styleSheet())
        self.right_button.setStyleSheet(self.styleSheet())
        self.set_x_button.setStyleSheet(self.styleSheet())
        self.set_y_button.setStyleSheet(self.styleSheet())
        self.reset_button.setStyleSheet(self.styleSheet())
        self.stop_button.setStyleSheet(self.styleSheet())
        self.layer_preview.setStyleSheet(self.styleSheet())
        self.layer_combobox.setStyleSheet(self.styleSheet())
      


    def add_movement_controls(self, layout):
        move_frame = QFrame()
        move_layout = QVBoxLayout(move_frame)
        
        # Add arrow buttons or joystick-like controls for manual movement
        # For example, add a button for moving in the +X direction
        move_x_plus = QPushButton("Move X+")
        move_x_plus.clicked.connect(lambda: self.manual_move("X+"))
        move_layout.addWidget(move_x_plus)

        layout.addWidget(move_frame)
    

    def manual_move(self, direction):
        # Placeholder for function to send manual movement commands
        print(f"Move in direction: {direction}")
        # Send the command to the Arduino, similar to send_to_arduino method

    def set_home(self):
        # Placeholder for function to set home position
        pass

    def load_excel_file(self):
        # This line uses the QFileDialog.getOpenFileName method to open the file dialog
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)")
        
        if file_name:
            self.excel_file = file_name
            workbook = openpyxl.load_workbook(self.excel_file)
            self.sheet_names = workbook.sheetnames
            self.layer_combobox.clear()
            self.layer_combobox.addItems(self.sheet_names)  # Use addItems to populate QComboBox
            print(f"Loaded Excel file: {self.excel_file}")


    def process_layer(self):
        selected_layer = self.layer_combobox.currentText()
        if not selected_layer or selected_layer == "Select Layer":
            QMessageBox.warning(self, "Layer Selection", "Please select a layer to process.")
            return
        
        workbook = openpyxl.load_workbook(self.excel_file)
        sheet = workbook[selected_layer]

        # Assuming the first 20x20 area corresponds to the grid
        commands = []
        for row in range(1, 21):  # Adjust range if your grid differs
            for col in range(1, 21):
                cell = sheet.cell(row, col)
                # Check if the cell's fill color is not white (or not transparent)
                # Assuming that default grid cells have no fill (white/transparent)
                if cell.fill.start_color.index != '00000000':  # '00000000' corresponds to no fill in openpyxl
                    commands.append(f"{col-1},{row-1};")  # Subtracting 1 to start coordinates at 0

        # Send commands to Arduino
        self.send_to_arduino(commands)
        
    def send_to_arduino(self, commands):
        if self.ser and self.ser.isOpen():
            for command in commands:
                print(f"Sending: {command}")
                self.ser.write(command.encode() + b'\n')  # Ensure newline is added if your Arduino sketch expects it
                self.ser.flush()  # Ensure data is sent immediately


def main():
    app = QApplication(sys.argv)
    gui = IndustrialCncGui()
    gui.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()

