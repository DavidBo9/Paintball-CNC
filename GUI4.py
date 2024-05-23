import sys
import serial
import time
import openpyxl
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel,
                             QVBoxLayout, QHBoxLayout, QWidget, QFrame, QComboBox, QFileDialog, QMessageBox, QGridLayout)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
from openpyxl.styles import PatternFill

class IndustrialCncGui(QMainWindow):
    
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.init_serial('/dev/cu.usbserial-10', 115200)  # Update the port
        self.is_processing = False  # Flag to control processing state
        
    def init_ui(self):
        self.setWindowTitle('Industrial Paintball CNC Controller')
        self.setGeometry(100, 100, 1280, 720)
        self.setup_ui_elements()
        self.style_interface()
        
    def init_serial(self, port, baud_rate):
        try:
            self.ser = serial.Serial(port, baud_rate)
            time.sleep(2)  # Allow time for the connection to initialize
            print(f"Connected to the Arduino on {port}.")
        except serial.SerialException as e:
            QMessageBox.critical(self, "Serial Connection Error", f"Failed to connect to the Arduino: {e}")
            self.ser = None

    def setup_ui_elements(self):
        self.setWindowTitle('Industrial Paintball CNC Controller')
        self.setGeometry(100, 100, 1280, 720)

        main_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        self.load_button = QPushButton('Load Excel File')
        self.load_button.setFixedSize(200, 40)
        self.layer_combobox = QComboBox()
        self.layer_preview = QLabel('Layer Preview')
        self.layer_preview.setFixedSize(300, 300)
        self.layer_preview.setStyleSheet("background-color: gray")
        self.process_button = QPushButton('Process Layer')
        self.process_button.setFixedSize(200, 40)
        self.process_button.setStyleSheet("background-color: green; color: white")

        left_layout.addWidget(self.load_button)
        left_layout.addWidget(self.layer_combobox)
        left_layout.addWidget(self.layer_preview)
        left_layout.addWidget(self.process_button)

        self.create_cnc_control_buttons(right_layout)

        main_layout.addLayout(left_layout)
        main_layout.addLayout(right_layout)

        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        self.load_button.clicked.connect(self.load_excel_file)
        self.process_button.clicked.connect(self.process_layer)

    def create_cnc_control_buttons(self, layout):
        self.shoot_button = QPushButton('SHOOT')
        self.shoot_button.clicked.connect(lambda: self.send_gcode_command("G91\nG0 F1000"))
        self.home_button = QPushButton('HOME')
        self.home_button.clicked.connect(self.move_home)

        self.stop_button = QPushButton('STOP')
        self.stop_button.clicked.connect(self.stop_processing)
        layout.addWidget(self.stop_button)  # Example command, adjust as needed
        self.up_button = QPushButton('UP')
        self.up_button.clicked.connect(lambda: self.send_gcode_command("G91\nG0 Y60 Z60"))
        self.down_button = QPushButton('DOWN')
        self.down_button.clicked.connect(lambda: self.send_gcode_command("G91\nG0 Y-60 Z-60"))
        self.left_button = QPushButton('LEFT')
        self.left_button.clicked.connect(lambda: self.send_gcode_command("G91\nG0 X-60"))
        self.right_button = QPushButton('RIGHT')
        self.right_button.clicked.connect(lambda: self.send_gcode_command("G91\nG0 X60"))

        layout.addWidget(self.shoot_button)
        layout.addWidget(self.home_button)
        layout.addWidget(self.up_button)
        layout.addWidget(self.down_button)
        layout.addWidget(self.left_button)
        layout.addWidget(self.right_button)

    def send_gcode_command(self, command, delay=1):
        if self.ser and self.ser.isOpen():
            full_command = f"{command}\n"
            print(f"Sending: {command}")
            self.ser.write(full_command.encode('utf-8'))
            self.ser.flush()
            response = self.ser.readline().decode('utf-8').strip()
            print(f"Received response: {response}")
            time.sleep(delay)  # Delay after sending the command
        else:
            QMessageBox.warning(self, "Connection Error", "Serial connection is not established.")

    def mock_shoot_action(self):
        """Simulate shooting action."""
        print("Simulating SHOOT action.")
        # Use a command suitable for your setup; here's a placeholder
        self.send_gcode_command("M300 S30", 5)  # Adjust as needed# Simulate the shoot action with a command
    
    def move_home(self):
        """Move the CNC to the origin (0,0) and reset the Z-axis."""
        self.send_gcode_command("G90\nG0 X0 Y0 Z0", 1)  # Use G90 to ensure absolute positioning
        print("Moved to HOME position.")

    def stop_processing(self):
        """Stop the current processing."""
        self.is_processing = False

    def load_excel_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)")
        if file_name:
            self.excel_file = file_name
            workbook = openpyxl.load_workbook(file_name)
            sheet_names = workbook.sheetnames
            self.layer_combobox.clear()
            self.layer_combobox.addItems(sheet_names)
   
    def process_layer(self):
        self.move_home()  # Ensure starting from the origin for each layer
        self.is_processing = True

        selected_layer = self.layer_combobox.currentText()
        if not selected_layer or self.ser is None or not self.ser.isOpen():
            QMessageBox.warning(self, "Error", "Please select a layer and ensure the device is connected.")
            return

        workbook = openpyxl.load_workbook(self.excel_file, read_only=True)
        sheet = workbook[selected_layer]
        for row in range(1, 50):
            for col in range(1, 50):
                if not self.is_processing:  # Allows for stopping the loop
                    return
                cell = sheet.cell(row, col)
                # Safely check for cell's fill color
                fill_color = cell.fill.start_color.index if cell.fill and cell.fill.start_color else '00000000'
                if fill_color == '00000000':  # Skip cells with default fill color
                    continue
                else:  # Cells with a fill color other than '00000000'
                    # Use absolute positioning for each move
                    self.send_gcode_command(f"G90\nG0 X{col-1} Y{row-1} Z{row-1}", 1)
                    self.mock_shoot_action()  # Simulate shooting



    def style_interface(self):
        self.setStyleSheet("""
        QMainWindow {
            background-color: #2c3e50;
        }
        QPushButton {
            background-color: #3498db;
            color: white;
            border-style: outset;
            border-width: 2px;
            border-radius: 10px;
            border-color: beige;
            font: bold 14px;
            padding: 6px;
            min-width: 10em;
            margin: 4px;
        }
        QLabel {
            background-color: #95a5a6;
            color: white;
            font: bold 14px;
            border-radius: 5px;
            padding: 6px;
            margin: 4px;
        }
        QComboBox {
            background-color: #34495e;
            color: white;
            border-radius: 5px;
            padding: 6px;
            margin: 4px;
            min-width: 10em;
        }
        """)

def main():
    app = QApplication(sys.argv)
    gui = IndustrialCncGui()
    gui.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
