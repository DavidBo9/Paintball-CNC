import tkinter as tk
from tkinter import filedialog, colorchooser, messagebox
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Function to convert an RGB value into an Excel fill color
def rgb_to_fill(rgb):
    return PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')

# Function to map image colors to a predefined palette
def map_colors(img, palette):
    # Convert the image to RGB if it's not already in RGB mode
    img = img.convert('RGB')
    
    # Create a new image with the same size as the original to apply the palette mapping
    new_img = Image.new('RGB', img.size)
    
    for x in range(img.size[0]):
        for y in range(img.size[1]):
            # Get the current pixel color
            current_color = img.getpixel((x, y))
            # Find the nearest color from the palette
            nearest_color = min(palette, key=lambda color: sum((s - q) ** 2 for s, q in zip(color, current_color)))
            # Set the pixel to the nearest color
            new_img.putpixel((x, y), nearest_color)
    
    return new_img

# GUI Functions
def upload_image():
    global img_path
    img_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.jpeg *.png")])
    if img_path:
        image_label.config(text="Image Loaded: " + img_path.split('/')[-1])

def choose_color(entry_widget):
    color_code = colorchooser.askcolor(title="Choose color")
    if color_code[1]:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, color_code[1])

def add_color_entry():
    color_entry = tk.Entry(frame, width=7)
    color_entry.pack(side=tk.LEFT)
    color_picker_btn = tk.Button(frame, text="Choose Color", command=lambda: choose_color(color_entry))
    color_picker_btn.pack(side=tk.LEFT)
    color_entries.append(color_entry)

def generate_pixel_art():
    try:
        # Retrieve the selected colors and dimensions from the GUI
        color_entries_hex = [entry.get() for entry in color_entries if entry.get() != '']
        colors_rgb = [tuple(int(hex_color[i:i+2], 16) for i in (1, 3, 5)) for hex_color in color_entries_hex]

        # Load and process the image, ensuring no alpha channel
        img = Image.open(img_path).convert('RGB')
        max_size = (int(width_entry.get()), int(height_entry.get()))
        img = img.resize(max_size, Image.Resampling.NEAREST)
        img_mapped = map_colors(img, colors_rgb)

        # Create a new Excel workbook for the color-mapped pixel art
        wb_mapped_palette = openpyxl.Workbook()
        complete_sheet = wb_mapped_palette.active
        complete_sheet.title = "Complete Image"

        # Fill in the complete image sheet with all layers combined
        for x in range(1, img_mapped.size[0] + 1):
            for y in range(1, img_mapped.size[1] + 1):
                pixel = img_mapped.getpixel((x - 1, y - 1))
                hex_color = ''.join([f'{value:02X}' for value in pixel])
                cell = complete_sheet.cell(row=y, column=x)
                cell.fill = rgb_to_fill(hex_color)
                complete_sheet.column_dimensions[get_column_letter(x)].width = 3
                complete_sheet.row_dimensions[y].height = 20

        # Generate a layer for each color
        for color in colors_rgb:
            hex_color = ''.join([f'{value:02X}' for value in color])
            ws = wb_mapped_palette.create_sheet(title=hex_color)
            for x in range(1, img_mapped.size[0] + 1):
                for y in range(1, img_mapped.size[1] + 1):
                    pixel = img_mapped.getpixel((x - 1, y - 1))
                    if pixel == color:  # Only fill the cell if it matches the layer's color
                        cell = ws.cell(row=y, column=x)
                        cell.fill = rgb_to_fill(hex_color)
                    ws.column_dimensions[get_column_letter(x)].width = 3
                    ws.row_dimensions[y].height = 20
        
        # Save the color-mapped Excel file
        output_file_path_mapped_palette = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                                      filetypes=[("Excel files", "*.xlsx")])
        if output_file_path_mapped_palette:
            wb_mapped_palette.save(output_file_path_mapped_palette)
            messagebox.showinfo("Success", "Pixel Art saved as " + output_file_path_mapped_palette)
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Set up the main application window
app = tk.Tk()
app.title("Pixel Art Generator")

# Layout
frame = tk.Frame(app)
frame.pack(padx=10, pady=10)

image_label = tk.Label(frame, text="No image selected")
image_label.pack()

upload_btn = tk.Button(frame, text="Upload Image", command=upload_image)
upload_btn.pack()

# List to store color entry widgets
color_entries = []

# Add default three color entries
for _ in range(3):
    add_color_entry()

# Button to add more color entries
add_color_btn = tk.Button(frame, text="Add Another Color", command=add_color_entry)
add_color_btn.pack()

# Entry for dimensions
width_label = tk.Label(frame, text="Width:")
width_label.pack(side=tk.LEFT)
width_entry = tk.Entry(frame, width=5)
width_entry.pack(side=tk.LEFT)
width_entry.insert(0, "50")  # Default value

height_label = tk.Label(frame, text="Height:")
height_label.pack(side=tk.LEFT)
height_entry = tk.Entry(frame, width=5)
height_entry.pack(side=tk.LEFT)
height_entry.insert(0, "50")  # Default value

# Button to generate pixel art
generate_btn = tk.Button(frame, text="Generate Pixel Art", command=generate_pixel_art)
generate_btn.pack()

app.mainloop()
