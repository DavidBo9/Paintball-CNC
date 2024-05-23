import sys
import serial
import openpyxl
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel,
                             QVBoxLayout, QHBoxLayout, QWidget, QComboBox,
                             QFileDialog, QMessageBox, QGridLayout)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
from openpyxl.styles import PatternFill

class IndustrialCncGui(QMainWindow):

    def __init__(self):
        super().__init__()
        self.ser = self.init_serial('/dev/cu.usbserial-10', 115200)  # Adjust the port and baud rate according to your setup
        self.init_ui()

    def init_serial(self, port, baud_rate):
        try:
            ser = serial.Serial(port, baud_rate)
            print(f"Connected to the Arduino on {port}.")
            return ser
        except Exception as e:
            QMessageBox.critical(self, "Serial Connection Error", f"Failed to connect to the Arduino: {e}")
            return None

    def init_ui(self):
        self.setWindowTitle('Industrial Paintball CNC Controller')
        self.setGeometry(100, 100, 1280, 720)
        self.setup_ui_elements()
        self.style_interface()

    def setup_ui_elements(self):
        main_layout = QHBoxLayout()
        self.create_control_panel(main_layout)
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

    def create_control_panel(self, main_layout):
        left_layout = self.create_excel_processing_layout()
        right_layout = self.create_cnc_control_layout()
        main_layout.addLayout(left_layout)
        main_layout.addLayout(right_layout)

    def create_excel_processing_layout(self):
        left_layout = QVBoxLayout()
        self.load_button = QPushButton('Load Excel File')
        self.load_button.setFixedSize(200, 40)
        self.layer_combobox = QComboBox()
        self.layer_preview = QLabel('Layer Preview')
        self.layer_preview.setFixedSize(300, 300)
        self.process_button = QPushButton('Process Layer')
        self.process_button.setFixedSize(200, 40)
        left_layout.addWidget(self.load_button)
        left_layout.addWidget(self.layer_combobox)
        left_layout.addWidget(self.layer_preview)
        left_layout.addWidget(self.process_button)
        return left_layout

    def create_cnc_control_layout(self):
        right_layout = QVBoxLayout()
        self.add_movement_controls(right_layout)
        self.stop_button = QPushButton('STOP', self)
        self.stop_button.setFixedSize(200, 100)
        right_layout.addWidget(self.stop_button)
        return right_layout

    def add_movement_controls(self, layout):
        move_grid = QGridLayout()
        # Define movement and action buttons
        self.home_button = QPushButton('HOME', self)
        self.up_button = QPushButton('UP', self)
        self.down_button = QPushButton('DOWN', self)
        self.left_button = QPushButton('LEFT', self)
        self.right_button = QPushButton('RIGHT', self)
        self.set_x_button = QPushButton('SET X-AXIS', self)
        self.set_y_button = QPushButton('SET Y-AXIS', self)
        self.reset_button = QPushButton('RESET', self)
        # Add buttons to the grid
        positions = [(0, 1, self.home_button), (1, 0, self.left_button), (1, 1, self.up_button),
                     (1, 2, self.right_button), (2, 1, self.down_button)]
        for pos in positions:
            move_grid.addWidget(pos[2], pos[0], pos[1])
        layout.addLayout(move_grid)
        layout.addWidget(self.set_x_button)
        layout.addWidget(self.set_y_button)
        layout.addWidget(self.reset_button)
        # Connect buttons to dummy functions for demonstration
        self.home_button.clicked.connect(lambda: self.dummy_function("HOME"))
        self.up_button.clicked.connect(lambda: self.dummy_function("UP"))
        # Repeat for other buttons as necessary

    def dummy_function(self, command):
        print(f"Command: {command}")

    def send_gcode_command(self, command):
        if self.ser and self.ser.isOpen():
            self.ser.write(f"{command}\n".encode('utf-8'))
            self.ser.flush()
            print(f"Sent: {command}")

    def load_excel_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)")
        if file_name:
            self.excel_file = file_name
            workbook = openpyxl.load_workbook(self.excel_file)
            self.sheet_names = workbook.sheetnames
            self.layer_combobox.clear()
            self.layer_combobox.addItems(self.sheet_names)

    def style_interface(self):
        self.setStyleSheet("""
        QMainWindow {
            background-color: #2c3e50;
        }
        QPushButton {
            background-color: #3498db;
            color: white;
            border-style: outset;
            border-width: 2px;
            border-radius: 10px;
            border-color: beige;
            font: bold 14px;
            padding: 6px;
            min-width: 10em;
            margin: 4px;
        }
        QPushButton#emergencyStopButton {
            background-color: #e74c3c;
        }
        QPushButton#processButton {
            background-color: #2ecc71;
        }
        QLabel {
            background-color: #95a5a6;
            color: white;
            font: bold 14px;
            border-radius: 5px;
            padding: 6px;
            margin: 4px;
        }
        QComboBox {
            background-color: #34495e;
            color: white;
            border-radius: 5px;
            padding: 6px;
            margin: 4px;
            min-width: 10em;
        }
        QSlider::groove:horizontal {
            border: 1px solid #bbb;
            background: white;
            height: 10px;
            border-radius: 4px;
        }
        QSlider::handle:horizontal {
            background: #3498db;
            border: 1px solid #777;
            width: 18px;
            margin: -2px 0;
            border-radius: 4px;
        }
        QSlider::add-page:horizontal {
            background: #bbb;
        }
        QSlider::sub-page:horizontal {
            background: #3498db;
        }
        """) # Your existing styles
               # Apply the button styles using objectName to distinguish between different buttons
        self.load_button.setObjectName('loadButton')
        self.process_button.setObjectName('processButton')
        self.shoot_button.setObjectName('shootButton')
        self.home_button.setObjectName('homeButton')
        self.up_button.setObjectName('upButton')
        self.down_button.setObjectName('downButton')
        self.left_button.setObjectName('leftButton')
        self.right_button.setObjectName('rightButton')
        self.set_x_button.setObjectName('setXButton')
        self.set_y_button.setObjectName('setYButton')
        self.reset_button.setObjectName('resetButton')
        self.stop_button.setObjectName('emergencyStopButton')

        # Reapply styles to update the widgets with the new object names
        self.load_button.setStyleSheet(self.styleSheet())
        self.process_button.setStyleSheet(self.styleSheet())
        self.shoot_button.setStyleSheet(self.styleSheet())
        self.home_button.setStyleSheet(self.styleSheet())
        self.up_button.setStyleSheet(self.styleSheet())
        self.down_button.setStyleSheet(self.styleSheet())
        self.left_button.setStyleSheet(self.styleSheet())
        self.right_button.setStyleSheet(self.styleSheet())
        self.set_x_button.setStyleSheet(self.styleSheet())
        self.set_y_button.setStyleSheet(self.styleSheet())
        self.reset_button.setStyleSheet(self.styleSheet())
        self.stop_button.setStyleSheet(self.styleSheet())
        self.layer_preview.setStyleSheet(self.styleSheet())
        self.layer_combobox.setStyleSheet(self.styleSheet())

def main():
    app = QApplication(sys.argv)
    gui = IndustrialCncGui()
    gui.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
